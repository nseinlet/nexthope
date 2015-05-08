# -*- coding: utf-8 -*-

from openerp import models, fields, api
import openpyxl
import datetime
import base64
import tempfile
import shutil

class XlsMoves(models.TransientModel):
    _name = 'nexthope_xls.xls_moves'

    state = fields.Char(selection=(('draft','draft'),('done','done')), default='draft')
    date_from = fields.Date('From')
    date_to = fields.Date('To')
    name = fields.Char()
    file = fields.Binary(readonly=True)
    
    @api.multi
    def do_export(self):
        tmpdir = tempfile.mkdtemp()
        tmpdir = tmpdir.rstrip('/')
        
        wb = openpyxl.Workbook()
        ws = wb.active

        line = 0
        am_model = self.env['account.move']
        moves = am_model.search([('date', '>=', self.date_from), ('date', '<=', self.date_to)])
        for move in moves:
            line = line + 1
            ws['A%s' % line] = move.ref
            for moveline in move.line_id:
                line = line + 1
                ws['B%s' % line] = moveline.account_id.name
                ws['C%s' % line] = moveline.debit
                ws['D%s' % line] = moveline.credit
            
        
        wb.save("%s/moves.xlsx" % tmpdir)
        
        with open("%s/moves.xlsx" % tmpdir, "rb") as xls_file:
            encoded_string = base64.b64encode(xls_file.read())
        
        self.name="moves.xlsx"
        self.file = encoded_string
        self.state = 'done'
        
        shutil.rmtree(tmpdir)
        
        return {
            'name': 'Download file',
            'context': self._context,
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'nexthope_xls.xls_moves',
            'type': 'ir.actions.act_window',
            'target': 'new',
            'res_id': self[0].id,
        }
        
        
