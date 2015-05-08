# -*- coding: utf-8 -*-

from openerp import models, fields, api
import openpyxl
import datetime
import base64

class XlsExport(models.TransientModel):
    _name = 'openacademy.xlsexport'

    state = fields.Char(selection=(('draft','draft'),('done','done')), default='draft')
    name = fields.Char()
    file = fields.Binary(readonly=True)
    
    @api.multi
    def do_export(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 42
        ws.append([1, 2, 3])
        ws['A2'] = datetime.datetime.now()
        wb.save("sample.xlsx")
        
        with open("sample.xlsx", "rb") as xls_file:
            encoded_string = base64.b64encode(xls_file.read())
        
        self.name="monfichier.xlsx"
        self.file = encoded_string
        self.state = 'done'
        
        return {
            'name': 'Download excel',
            'context': self._context,
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'openacademy.xlsexport',
            'type': 'ir.actions.act_window',
            'target': 'new',
            'res_id': self[0].id,
        }
        
        